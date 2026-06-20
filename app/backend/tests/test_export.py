"""Tests for the SQLite export module."""

import io
from datetime import date

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from sqlmodel import SQLModel

from models import Sighting
from routers.export import EXPORT_TABLES, RECORD_HEADERS

# Tables that intentionally should NOT be in the export
INTERNAL_TABLES = {"organisation"}


def _make_sighting(
    db_session: Session,
    survey_id: int,
    species_id: int,
    count: int = 1,
    location_id: int = None,
) -> Sighting:
    sighting = Sighting(
        survey_id=survey_id,
        species_id=species_id,
        count=count,
        location_id=location_id,
    )
    db_session.add(sighting)
    db_session.commit()
    db_session.refresh(sighting)
    return sighting


def _read_rows(response) -> list[list]:
    """Parse an xlsx response into a list of rows (including the header row)."""
    workbook = load_workbook(io.BytesIO(response.content))
    return [list(row) for row in workbook.active.iter_rows(values_only=True)]


class TestExportCompleteness:
    """Ensure the export table list stays in sync with the database schema."""

    def test_all_tables_accounted_for(self):
        """Fail if a SQLModel table exists that is neither exported nor explicitly internal."""
        all_tables = set(SQLModel.metadata.tables.keys())
        exported = {t["table"] for t in EXPORT_TABLES}
        accounted = exported | INTERNAL_TABLES
        unaccounted = all_tables - accounted

        assert not unaccounted, (
            f"Tables {unaccounted} exist in SQLModel metadata but are not in "
            f"EXPORT_TABLES or INTERNAL_TABLES in routers/export.py. "
            f"Add them to one or the other."
        )


class TestRecordsExportBySurveyType:
    """Tests for GET /api/export/records/by-survey-type/{id}"""

    def test_exports_records_with_expected_columns(
        self,
        client: TestClient,
        auth_headers: dict,
        db_session: Session,
        create_survey_type,
        create_survey,
        create_species,
        create_location,
    ):
        survey_type = create_survey_type(name="Dragonfly Transect")
        location = create_location(name="North Meadow")
        species = create_species(
            name="Emperor Dragonfly", scientific_name="Anax imperator"
        )
        survey = create_survey(
            survey_date=date(2026, 6, 1),
            survey_type_id=survey_type.id,
            location_id=location.id,
        )
        _make_sighting(db_session, survey.id, species.id, count=3)

        response = client.get(
            f"/api/export/records/by-survey-type/{survey_type.id}",
            headers=auth_headers,
        )

        assert response.status_code == 200
        assert "spreadsheetml" in response.headers["content-type"]
        # Filename like "survey_dragonfly_transect_<datetime>.xlsx"
        disposition = response.headers["content-disposition"]
        assert "filename=\"survey_dragonfly_transect_" in disposition
        assert disposition.endswith('.xlsx"')

        rows = _read_rows(response)
        assert rows[0] == RECORD_HEADERS
        assert len(rows) == 2
        common_name, species_name, count, day, loc = rows[1]
        assert common_name == "Emperor Dragonfly"
        assert species_name == "Anax imperator"
        assert count == 3
        assert str(day).startswith("2026-06-01")
        assert loc == "North Meadow"

    def test_sighting_location_takes_precedence_over_survey(
        self,
        client: TestClient,
        auth_headers: dict,
        db_session: Session,
        create_survey_type,
        create_survey,
        create_species,
        create_location,
    ):
        survey_type = create_survey_type()
        survey_location = create_location(name="Survey Location")
        sighting_location = create_location(name="Sighting Location")
        species = create_species()
        survey = create_survey(
            survey_type_id=survey_type.id, location_id=survey_location.id
        )
        _make_sighting(
            db_session, survey.id, species.id, location_id=sighting_location.id
        )

        response = client.get(
            f"/api/export/records/by-survey-type/{survey_type.id}",
            headers=auth_headers,
        )

        rows = _read_rows(response)
        assert rows[1][4] == "Sighting Location"

    def test_only_includes_matching_survey_type(
        self,
        client: TestClient,
        auth_headers: dict,
        db_session: Session,
        create_survey_type,
        create_survey,
        create_species,
    ):
        wanted = create_survey_type(name="Wanted")
        other = create_survey_type(name="Other")
        species = create_species()
        wanted_survey = create_survey(survey_type_id=wanted.id)
        other_survey = create_survey(survey_type_id=other.id)
        _make_sighting(db_session, wanted_survey.id, species.id, count=5)
        _make_sighting(db_session, other_survey.id, species.id, count=9)

        response = client.get(
            f"/api/export/records/by-survey-type/{wanted.id}", headers=auth_headers
        )

        rows = _read_rows(response)
        assert len(rows) == 2  # header + the single matching sighting
        assert rows[1][2] == 5

    def test_unknown_survey_type_returns_404(
        self, client: TestClient, auth_headers: dict
    ):
        response = client.get(
            "/api/export/records/by-survey-type/999999", headers=auth_headers
        )
        assert response.status_code == 404

    def test_requires_authentication(self, client: TestClient):
        response = client.get("/api/export/records/by-survey-type/1")
        assert response.status_code == 401


class TestRecordsExportBySpeciesType:
    """Tests for GET /api/export/records/by-species-type/{id}"""

    def test_includes_sightings_across_survey_types(
        self,
        client: TestClient,
        auth_headers: dict,
        db_session: Session,
        create_survey_type,
        create_survey,
        create_species,
        create_species_type,
    ):
        dragonfly = create_species_type(name="dragonfly", display_name="Dragonfly")
        butterfly_species = create_species(
            name="Peacock", scientific_name="Aglais io", species_type="butterfly"
        )
        dragonfly_species = create_species(
            name="Emperor", scientific_name="Anax imperator", species_type="dragonfly"
        )
        transect = create_survey_type(name="Transect")
        adhoc = create_survey_type(name="Ad-hoc")
        transect_survey = create_survey(survey_type_id=transect.id)
        adhoc_survey = create_survey(survey_type_id=adhoc.id)
        # Same dragonfly species seen in two different survey types
        _make_sighting(db_session, transect_survey.id, dragonfly_species.id, count=2)
        _make_sighting(db_session, adhoc_survey.id, dragonfly_species.id, count=4)
        # A butterfly sighting that must be excluded
        _make_sighting(db_session, transect_survey.id, butterfly_species.id, count=7)

        response = client.get(
            f"/api/export/records/by-species-type/{dragonfly.id}",
            headers=auth_headers,
        )

        assert response.status_code == 200
        rows = _read_rows(response)
        assert len(rows) == 3  # header + two dragonfly sightings, no butterfly
        counts = sorted(row[2] for row in rows[1:])
        assert counts == [2, 4]

    def test_unknown_species_type_returns_404(
        self, client: TestClient, auth_headers: dict
    ):
        response = client.get(
            "/api/export/records/by-species-type/999999", headers=auth_headers
        )
        assert response.status_code == 404


class TestExportableTypeLists:
    """The list endpoints only surface types that have at least one record."""

    def test_survey_types_list_excludes_empty(
        self,
        client: TestClient,
        auth_headers: dict,
        db_session: Session,
        create_survey_type,
        create_survey,
        create_species,
    ):
        with_records = create_survey_type(name="With Records")
        create_survey_type(name="Empty")  # no sightings -> excluded
        species = create_species()
        survey = create_survey(survey_type_id=with_records.id)
        _make_sighting(db_session, survey.id, species.id)

        response = client.get(
            "/api/export/records/survey-types", headers=auth_headers
        )

        assert response.status_code == 200
        names = [st["name"] for st in response.json()]
        assert names == ["With Records"]

    def test_species_types_list_excludes_empty(
        self,
        client: TestClient,
        auth_headers: dict,
        db_session: Session,
        create_survey_type,
        create_survey,
        create_species,
        create_species_type,
    ):
        # An "empty" species type with a species but no sightings.
        create_species_type(name="moth", display_name="Moth")
        create_species(name="A Moth", species_type="moth")
        # A species type that does have a sighting.
        butterfly = create_species(species_type="butterfly")
        survey = create_survey(survey_type_id=create_survey_type().id)
        _make_sighting(db_session, survey.id, butterfly.id)

        response = client.get(
            "/api/export/records/species-types", headers=auth_headers
        )

        assert response.status_code == 200
        display_names = [st["display_name"] for st in response.json()]
        assert "Moth" not in display_names
        assert "Butterfly" in display_names

    def test_lists_require_authentication(self, client: TestClient):
        assert client.get("/api/export/records/survey-types").status_code == 401
        assert client.get("/api/export/records/species-types").status_code == 401
